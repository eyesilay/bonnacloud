import os
import io
import json
import zipfile
import mimetypes
import urllib.parse
import requests
import openpyxl
import threading 
from typing import List, Optional
from datetime import datetime, timedelta

from fastapi import FastAPI, HTTPException, Depends, Security, UploadFile, File, BackgroundTasks, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text
from jose import JWTError, jwt
import bcrypt

from database import SessionLocal, UserDB, RoleDB, Base, engine

# =========================================================
# 🛡️ DATABASE SCHEMAS
# =========================================================
class RoleRequest(BaseModel):
    name: str
    allowedFolders: List[dict] = []
    isAdmin: Optional[bool] = False

class UserCreateRequest(BaseModel):
    name: str
    email: str
    password: str
    role: Optional[str] = "customer"
    role_id: Optional[int] = None
    company: Optional[str] = None
    isActive: bool = True

class UserUpdateRequest(BaseModel):
    name: str
    email: str
    password: Optional[str] = None
    role: Optional[str] = "customer"
    role_id: Optional[int] = None
    company: Optional[str] = None
    isActive: bool = True

class ZipDownloadRequest(BaseModel):
    paths: List[str]

class LoginRequest(BaseModel):
    email: str
    password: str

class CDNConfigUpdateRequest(BaseModel):
    storageName: str
    storagePassword: str
    pullZoneUrl: str
    region: str

# =========================================================
# INFRASTRUCTURE & SECURITY CONFIGURATION
# =========================================================
os.makedirs("/app/data", exist_ok=True)
Base.metadata.create_all(bind=engine)

SECRET_KEY = os.getenv("BONNA_JWT_SECRET", "bonna-secure-dynamic-node-crypto-key-2026")
ALGORITHM = "HS256"

security_bearer = HTTPBearer()

GLOBAL_FILE_INDEX = []
LAST_INDEX_TIME = None
CACHE_DURATION_DAYS = 7  
IS_INDEXING = False

def get_password_hash(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try: return bcrypt.checkpw(plain_password.encode('utf-8'), hashed_password.encode('utf-8'))
    except: return False

def create_access_token(data: dict):
    to_encode = data.copy()
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

def get_current_user(credentials: HTTPAuthorizationCredentials = Security(security_bearer), db: Session = Depends(get_db)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None: raise HTTPException(status_code=401, detail="Invalid session token.")
    except JWTError: raise HTTPException(status_code=401, detail="Session verification failed.")
    
    user = db.query(UserDB).filter(UserDB.email == email).first()
    if not user: raise HTTPException(status_code=401, detail="User not found in system.")
    if not user.is_active: raise HTTPException(status_code=403, detail="Your account has been suspended.")
    
    if user.is_superadmin:
        user.role = "superadmin"
    elif user.role_rel and user.role_rel.is_admin:
        user.role = "admin"
    return user

def init_db():
    try:
        with engine.connect() as conn:
            try: conn.execute(text("ALTER TABLE users ADD COLUMN failed_login_attempts INTEGER DEFAULT 0;")); conn.commit()
            except: pass
            try: conn.execute(text("ALTER TABLE users ADD COLUMN lockout_until DATETIME;")); conn.commit()
            except: pass
            try: conn.execute(text("ALTER TABLE users ADD COLUMN role_id INTEGER REFERENCES roles(id);")); conn.commit()
            except: pass
            try: conn.execute(text("ALTER TABLE users ADD COLUMN company VARCHAR;")); conn.commit()
            except: pass
            try: conn.execute(text("ALTER TABLE roles ADD COLUMN is_admin BOOLEAN DEFAULT 0;")); conn.commit()
            except: pass
            # Otomatik kolon koruması: SQLite veritabanına is_superadmin alanını enjekte eder
            try: conn.execute(text("ALTER TABLE users ADD COLUMN is_superadmin BOOLEAN DEFAULT 0;")); conn.commit()
            except: pass
    except Exception as e: print(f"Database sync alert: {e}")

    db = SessionLocal()
    try:
        admin_role = db.query(RoleDB).filter(RoleDB.name == "Administrator").first()
        if not admin_role:
            admin_role = RoleDB(name="Administrator", allowed_folders=[{"id": "", "name": "Full Server", "allowed": True}], is_admin=True)
            db.add(admin_role)
            db.commit()
            db.refresh(admin_role)

        admin = db.query(UserDB).filter(UserDB.role == "admin").first()
        if not admin:
            raw_password = os.getenv("BONNA_ADMIN_PASSWORD", "admin")
            new_admin = UserDB(name="System Administrator", email="admin@bonnacloud.com", password=get_password_hash(raw_password), role="admin", role_id=admin_role.id, is_active=True)
            db.add(new_admin)
            db.commit()
        elif admin and not admin.role_id:
            admin.role_id = admin_role.id
            db.commit()
    finally: db.close()

init_db()

app = FastAPI(title="Bonna Cloud Engine")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"], expose_headers=["Content-Disposition"])

CONFIG_FILE = "/app/data/cdn_config.json"
DEFAULT_CONFIG = {"BUNNY_STORAGE_NAME": "bonna-api-test", "BUNNY_STORAGE_PASSWORD": "password", "BUNNY_PULL_ZONE_URL": "url", "BUNNY_REGION": ""}

def load_cdn_config():
    if os.path.exists(CONFIG_FILE):
        try: return json.load(open(CONFIG_FILE, "r"))
        except: return DEFAULT_CONFIG
    return DEFAULT_CONFIG

def save_cdn_config(config_data):
    with open(CONFIG_FILE, "w") as f: json.dump(config_data, f)

def get_bunny_base_url():
    cfg = load_cdn_config()
    storage_name = cfg.get("BUNNY_STORAGE_NAME")
    region = cfg.get("BUNNY_REGION")
    region_prefix = f"{region}." if region else ""
    return f"https://{region_prefix}storage.bunnycdn.com/{storage_name}"

def clean_path(p: str) -> str:
    if not p: return ""
    return p.strip("/")

def background_index_crawl():
    global GLOBAL_FILE_INDEX, LAST_INDEX_TIME, IS_INDEXING
    if IS_INDEXING:
        return
    IS_INDEXING = True
    try:
        cfg = load_cdn_config()
        storage_name = cfg.get("BUNNY_STORAGE_NAME")
        if storage_name == "bonna-api-test" or not cfg.get("BUNNY_STORAGE_PASSWORD"):
            return
            
        base_url = get_bunny_base_url()
        headers = {"AccessKey": cfg.get("BUNNY_STORAGE_PASSWORD"), "accept": "application/json"}
        compiled_list = []
        queue = [""]
        folders_scanned = 0
        
        while queue and folders_scanned < 100000:
            current_path = queue.pop(0)
            folders_scanned += 1
            url = f"{base_url}/{urllib.parse.quote(current_path, safe='/')}/" if current_path else f"{base_url}/"
            try:
                resp = requests.get(url, headers=headers, timeout=10)
                if resp.status_code == 200:
                    for item in resp.json():
                        is_dir = item.get("IsDirectory", False)
                        name = item.get("ObjectName")
                        item_path = f"{current_path}/{name}" if current_path else name
                        mime_type = "application/vnd.google-apps.folder" if is_dir else (mimetypes.guess_type(name)[0] or "application/octet-stream")
                        
                        compiled_list.append({
                            "id": item_path,
                            "name": name,
                            "mimeType": mime_type,
                            "size": item.get("Length", 0),
                            "webViewLink": f"{cfg.get('BUNNY_PULL_ZONE_URL')}/{urllib.parse.quote(item_path, safe='/')}" if not is_dir else ""
                        })
                        if is_dir:
                            queue.append(item_path)
            except:
                pass
                
        GLOBAL_FILE_INDEX = compiled_list
        LAST_INDEX_TIME = datetime.utcnow()
        print(f"--- BonnaCloud İndeksleme Tamamlandı: {len(GLOBAL_FILE_INDEX)} nesne hafızaya alındı. ---")
    except Exception as e:
        print(f"İndeksleme sırasında hata oluştu: {e}")
    finally:
        IS_INDEXING = False

@app.on_event("startup")
def startup_event():
    t = threading.Thread(target=background_index_crawl)
    t.daemon = True
    t.start()

# ROLE API
@app.get("/api/roles")
def get_roles(db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    return [{"id": r.id, "name": r.name, "allowedFolders": r.allowed_folders or [], "isAdmin": r.is_admin} for r in db.query(RoleDB).all()]

@app.post("/api/roles")
def create_role(payload: RoleRequest, db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    if db.query(RoleDB).filter(RoleDB.name == payload.name).first(): raise HTTPException(status_code=400, detail="A role with this name already exists.")
    db.add(RoleDB(name=payload.name, allowed_folders=payload.allowedFolders, is_admin=payload.isAdmin))
    db.commit()
    return {"message": "Role successfully created."}

@app.put("/api/roles/{role_id}")
def update_role(role_id: int, payload: RoleRequest, db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    role = db.query(RoleDB).filter(RoleDB.id == role_id).first()
    if not role: raise HTTPException(status_code=404, detail="Role not found.")
    
    role.name = payload.name
    role.allowed_folders = payload.allowedFolders
    role.is_admin = payload.isAdmin
    
    db.query(UserDB).filter(UserDB.role_id == role_id, UserDB.is_superadmin == False).update({UserDB.role: "admin" if payload.isAdmin else "customer"})
    db.commit()
    return {"message": "Role successfully updated."}

@app.delete("/api/roles/{role_id}")
def delete_role(role_id: int, db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    role = db.query(RoleDB).filter(RoleDB.id == role_id).first()
    if not role: raise HTTPException(status_code=404, detail="Role not found.")
    db.query(UserDB).filter(UserDB.role_id == role_id, UserDB.is_superadmin == False).update({UserDB.role_id: None, UserDB.role: "customer"})
    db.delete(role)
    db.commit()
    return {"message": "Role successfully deleted."}

# USER API
@app.get("/api/users")
def get_users(db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    return [{"id": u.id, "name": u.name, "email": u.email, "role_id": u.role_id, "role_name": u.role_rel.name if u.role_rel else "No Role Assigned", "company": u.company, "isActive": u.is_active, "isSuperAdmin": u.is_superadmin} for u in db.query(UserDB).all()]

@app.post("/api/users")
def create_user(payload: UserCreateRequest, db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    if db.query(UserDB).filter(UserDB.email == payload.email).first(): raise HTTPException(status_code=400, detail="This email address is already in use.")
    
    text_role = "customer"
    if payload.role_id:
        r_obj = db.query(RoleDB).filter(RoleDB.id == payload.role_id).first()
        if r_obj and r_obj.is_admin:
            text_role = "admin"
            
    db.add(UserDB(name=payload.name, email=payload.email, password=get_password_hash(payload.password), role=text_role, role_id=payload.role_id, company=payload.company, is_active=payload.isActive, is_superadmin=False))
    db.commit()
    return {"message": "User successfully created."}

@app.put("/api/users/{user_id}")
def update_user(user_id: int, payload: UserUpdateRequest, db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    target_user = db.query(UserDB).filter(UserDB.id == user_id).first()
    if not target_user: raise HTTPException(status_code=404, detail="User not found.")
    
    # 👑 KORUMA: Normal admin Süper Admin olan bir kullanıcıyı düzenleyemez
    if target_user.is_superadmin and current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Süper Admin seviyesindeki kullanıcılara müdahale edemezsiniz.")
        
    text_role = "customer"
    if payload.role_id:
        r_obj = db.query(RoleDB).filter(RoleDB.id == payload.role_id).first()
        if r_obj and r_obj.is_admin:
            text_role = "admin"
            
    target_user.name = payload.name
    target_user.email = payload.email
    target_user.role_id = payload.role_id
    if not target_user.is_superadmin:
        target_user.role = text_role
    target_user.company = payload.company
    target_user.is_active = payload.isActive
    if payload.password: target_user.password = get_password_hash(payload.password)
    db.commit()
    return {"message": "User successfully updated."}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    target_user = db.query(UserDB).filter(UserDB.id == user_id).first()
    if not target_user: raise HTTPException(status_code=404, detail="User not found.")
    if target_user.id == current_user.id: raise HTTPException(status_code=400, detail="You cannot delete your own account.")
    
    # 👑 KORUMA: Normal admin Süper Admin kullanıcısını silemez
    if target_user.is_superadmin and current_user.role != "superadmin":
        raise HTTPException(status_code=403, detail="Süper Admin kullanıcıları silme yetkiniz yoktur.")
        
    db.delete(target_user)
    db.commit()
    return {"message": "User successfully deleted."}

@app.post("/api/users/bulk")
async def bulk_create_users(file: UploadFile = File(...), db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(content), data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in ws[1]]
        all_roles = db.query(RoleDB).all()
        def to_lower_safe(text): return str(text).replace('I', 'ı').replace('İ', 'i').strip().lower()
        added_count = 0
        updated_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(raw_headers, row))
            email, password, role_name, company = None, None, None, None
            for key, val in row_data.items():
                if val is None: continue
                if "mail" in key: email = val
                elif "pass" in key or "şifre" in key or "sifre" in key: password = val
                elif "role" in key or "rol" in key: role_name = val
                elif "company" in key or "şirket" in key or "sirket" in key: company = val
            if not email or not password: continue
            email = str(email).strip()
            password = str(password).strip()
            company_str = str(company).strip() if company is not None else None
            role_id = None
            if role_name is not None and str(role_name).strip() != "":
                target_role = to_lower_safe(role_name)
                for r in all_roles:
                    if to_lower_safe(r.name) == target_role: role_id = r.id; break
                if not role_id:
                    for r in all_roles:
                        if target_role in to_lower_safe(r.name): role_id = r.id; break
            
            is_admin_role = False
            if role_id:
                r_obj = db.query(RoleDB).filter(RoleDB.id == role_id).first()
                if r_obj and r_obj.is_admin:
                    is_admin_role = True
            text_role = "admin" if is_admin_role else "customer"

            existing = db.query(UserDB).filter(UserDB.email == email).first()
            if existing:
                if existing.is_superadmin and current_user.role != "superadmin":
                    continue # Normal admin Excel listesinden Süper Admini ezemez
                existing.role_id = role_id
                if not existing.is_superadmin: existing.role = text_role 
                if company_str: existing.company = company_str
                updated_count += 1
            else:
                db.add(UserDB(name=email.split('@')[0], email=email, password=get_password_hash(password), role=text_role, role_id=role_id, company=company_str, is_active=True, is_superadmin=False))
                added_count += 1
        db.commit()
        return {"message": f"System: {added_count} new users added, {updated_count} users successfully updated."}
    except Exception as e: raise HTTPException(status_code=400, detail=f"Excel processing error: {str(e)}")

@app.post("/api/login")
def login(payload: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(UserDB).filter(UserDB.email == payload.email).first()
    if not user or not user.is_active: raise HTTPException(status_code=401, detail="Login credentials verification failed.")
    if not verify_password(payload.password, user.password): raise HTTPException(status_code=401, detail="Login credentials verification failed.")
    
    role_text = user.role
    if user.is_superadmin:
        role_text = "superadmin"
    elif user.role_rel and user.role_rel.is_admin:
        role_text = "admin"
        
    token = create_access_token(data={"sub": user.email, "role": role_text})
    resolved_folders = user.role_rel.allowed_folders if user.role_rel else (user.allowed_folders if user.allowed_folders else [])
    
    return {
        "id": user.id, 
        "name": user.name or "System Administrator", 
        "email": user.email, 
        "role": role_text, 
        "role_name": user.role_rel.name if user.role_rel else ("Super Admin" if user.is_superadmin else "Customer"),
        "role_id": user.role_id, 
        "allowedFolders": resolved_folders, 
        "isActive": user.is_active, 
        "token": token
    }

@app.get("/api/cdn")
def get_cdn_settings(current_user: UserDB = Depends(get_current_user)):
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    cfg = load_cdn_config()
    return {"storageName": cfg.get("BUNNY_STORAGE_NAME"), "storagePassword": cfg.get("BUNNY_STORAGE_PASSWORD"), "pullZoneUrl": cfg.get("BUNNY_PULL_ZONE_URL"), "region": cfg.get("BUNNY_REGION")}

@app.post("/api/cdn")
def update_cdn_settings(payload: CDNConfigUpdateRequest, background_tasks: BackgroundTasks, current_user: UserDB = Depends(get_current_user)):
    global LAST_INDEX_TIME, GLOBAL_FILE_INDEX
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    save_cdn_config({"BUNNY_STORAGE_NAME": payload.storageName, "BUNNY_STORAGE_PASSWORD": payload.storagePassword, "BUNNY_PULL_ZONE_URL": payload.pullZoneUrl, "BUNNY_REGION": payload.region})
    LAST_INDEX_TIME = None 
    GLOBAL_FILE_INDEX = []
    background_tasks.add_task(background_index_crawl)
    return {"message": "CDN configuration node successfully updated."}

@app.post("/api/cdn/index")
def trigger_manual_indexing(background_tasks: BackgroundTasks, current_user: UserDB = Depends(get_current_user)):
    global IS_INDEXING
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    if IS_INDEXING: return {"message": "Indexing is already running in background.", "status": "running"}
    background_tasks.add_task(background_index_crawl)
    return {"message": "Manual search index task successfully deployed.", "status": "started"}

@app.get("/api/cdn/index-status")
def get_index_status(current_user: UserDB = Depends(get_current_user)):
    global LAST_INDEX_TIME, GLOBAL_FILE_INDEX, IS_INDEXING
    if current_user.role not in ["admin", "superadmin"]: raise HTTPException(status_code=403, detail="Permission denied.")
    return {
        "is_indexing": IS_INDEXING,
        "last_index_time": LAST_INDEX_TIME.isoformat() if LAST_INDEX_TIME else None,
        "total_items": len(GLOBAL_FILE_INDEX)
    }

def check_hierarchical_permission(item_path: str, allowed_folders: list) -> bool:
    path_clean = clean_path(item_path)
    if not path_clean: return True
    allowed_map = {clean_path(f.get("id", "")): f.get("allowed", True) for f in allowed_folders}
    if path_clean in allowed_map: return allowed_map[path_clean] is True
    parts = path_clean.split('/')
    for i in range(len(parts) - 1, 0, -1):
        if '/'.join(parts[:i]) in allowed_map: return allowed_map['/'.join(parts[:i])] is True
    return False

@app.get("/files")
def get_files(path: str = "", db: Session = Depends(get_db), current_user: UserDB = Depends(get_current_user)):
    clean_p = clean_path(path)
    cfg = load_cdn_config()
    allowed_folders = current_user.role_rel.allowed_folders if current_user.role_rel else []
    
    if current_user.role in ["customer", "müşteri"]:
        if not allowed_folders: allowed_folders = current_user.allowed_folders or []
        if not clean_p:
            allowed_entries = [f for f in allowed_folders if f.get("allowed", True) is True]
            result = []
            for f in allowed_entries:
                m_type = f.get("mimeType", "application/vnd.google-apps.folder")
                is_dir = (m_type == "application/vnd.google-apps.folder")
                item_path = clean_path(f.get("id", ""))
                result.append({"id": item_path, "name": f.get("name"), "mimeType": m_type, "size": f.get("size", 0), "webViewLink": "" if is_dir else f"{cfg.get('BUNNY_PULL_ZONE_URL')}/{urllib.parse.quote(item_path, safe='/')}"})
            return {"files": result}
        if not check_hierarchical_permission(clean_p, allowed_folders): return {"error": "Access denied to this storage layer.", "files": []}

    quoted_p = urllib.parse.quote(clean_p, safe='/')
    url = f"{get_bunny_base_url()}/{quoted_p}/" if quoted_p else f"{get_bunny_base_url()}/"
    try:
        resp = requests.get(url, headers={"AccessKey": cfg.get("BUNNY_STORAGE_PASSWORD"), "accept": "application/json"}, timeout=15)
        if resp.status_code != 200: return {"files": []}
    except: raise HTTPException(status_code=502, detail="CDN storage gateway link failure.")
    result = []
    for item in resp.json():
        is_dir = item.get("IsDirectory", False)
        name = item.get("ObjectName")
        item_path = f"{clean_p}/{name}" if clean_p else name
        if current_user.role in ["customer", "müşteri"] and not check_hierarchical_permission(item_path, allowed_folders): continue
        mime_type = "application/vnd.google-apps.folder" if is_dir else (mimetypes.guess_type(name)[0] or "application/octet-stream")
        result.append({"id": item_path, "name": name, "mimeType": mime_type, "size": item.get("Length", 0), "webViewLink": f"{cfg.get('BUNNY_PULL_ZONE_URL')}/{urllib.parse.quote(item_path, safe='/')}" if not is_dir else ""})
    return {"files": result}

@app.get("/api/search")
def search_files(query: str, background_tasks: BackgroundTasks, current_user: Session = Depends(get_db), current_user_obj: UserDB = Depends(get_current_user)):
    global GLOBAL_FILE_INDEX, LAST_INDEX_TIME, IS_INDEXING
    now = datetime.utcnow()
    if not LAST_INDEX_TIME or (now - LAST_INDEX_TIME) > timedelta(days=CACHE_DURATION_DAYS):
        if not IS_INDEXING: background_tasks.add_task(background_index_crawl)
            
    result_list = []
    query_lower = query.lower()
    allowed_folders = current_user_obj.role_rel.allowed_folders if current_user_obj.role_rel else []
    if not allowed_folders: allowed_folders = current_user_obj.allowed_folders or []

    for item in GLOBAL_FILE_INDEX:
        item_path = item["id"]
        name = item["name"]
        if current_user_obj.role in ["customer", "müşteri"]:
            if not check_hierarchical_permission(item_path, allowed_folders): continue
            if not any(item_path == clean_path(f.get("id", "")) or item_path.startswith(clean_path(f.get("id", "")) + "/") for f in allowed_folders): continue
                
        if query_lower in name.lower():
            result_list.append(item)
            
    return {"files": result_list}

@app.get("/download/file")
def download_file(path: str, current_user: UserDB = Depends(get_current_user)):
    cfg = load_cdn_config()
    url = f"{get_bunny_base_url()}/{urllib.parse.quote(clean_path(path), safe='/')}"
    try:
        resp = requests.get(url, headers={"AccessKey": cfg.get("BUNNY_STORAGE_PASSWORD")}, stream=True)
        if resp.status_code != 200: raise HTTPException(status_code=404, detail="Requested object not found.")
        return StreamingResponse(resp.iter_content(chunk_size=8192), media_type=resp.headers.get("Content-Type", "application/octet-stream"), headers={"Content-Disposition": f"attachment; filename={urllib.parse.quote(path.split('/')[-1])}"})
    except: raise HTTPException(status_code=502, detail="CDN network streaming error.")

@app.get("/api/view/pdf")
def view_pdf(path: str, token: str, request: Request, db: Session = Depends(get_db)):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email: str = payload.get("sub")
        if email is None: raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError: raise HTTPException(status_code=401, detail="Invalid token")
    
    user = db.query(UserDB).filter(UserDB.email == email).first()
    if not user or not user.is_active: raise HTTPException(status_code=401, detail="Unauthorized")
    
    allowed_folders = user.role_rel.allowed_folders if user.role_rel else (user.allowed_folders or [])
    if user.role in ["customer", "müşteri"]:
        if not check_hierarchical_permission(path, allowed_folders): raise HTTPException(status_code=403, detail="Forbidden")
            
    cfg = load_cdn_config()
    url = f"{get_bunny_base_url()}/{urllib.parse.quote(clean_path(path), safe='/')}"
    
    bunny_headers = {"AccessKey": cfg.get("BUNNY_STORAGE_PASSWORD")}
    client_range = request.headers.get("range")
    if client_range:
        bunny_headers["Range"] = client_range
        
    try:
        resp = requests.get(url, headers=bunny_headers, stream=True)
        if resp.status_code not in [200, 206]: 
            raise HTTPException(status_code=resp.status_code, detail="CDN stream failure")
        
        response_headers = {
            "Content-Type": "application/pdf",
            "Content-Disposition": "inline",
            "Accept-Ranges": "bytes",
            "X-Frame-Options": "ALLOWALL",
            "Content-Security-Policy": "frame-ancestors 'self' *",
            "Cache-Control": "public, max-age=3600"
        }
        
        if "Content-Range" in resp.headers:
            response_headers["Content-Range"] = resp.headers["Content-Range"]
        if "Content-Length" in resp.headers:
            response_headers["Content-Length"] = resp.headers["Content-Length"]
            
        return StreamingResponse(
            resp.iter_content(chunk_size=8192), 
            status_code=resp.status_code,
            headers=response_headers
        )
    except Exception as e: 
        raise HTTPException(status_code=502, detail=f"CDN proxy loop error: {str(e)}")

@app.post("/download/zip")
def download_zip(payload: ZipDownloadRequest, current_user: UserDB = Depends(get_current_user)):
    cfg = load_cdn_config()
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for path in payload.paths:
            url = f"{get_bunny_base_url()}/{urllib.parse.quote(clean_path(path), safe='/')}"
            try:
                resp = requests.get(url, headers={"AccessKey": cfg.get("BUNNY_STORAGE_PASSWORD")})
                if resp.status_code == 200: zip_file.writestr(path.split("/")[-1], resp.content)
            except: pass
    zip_buffer.seek(0)
    return StreamingResponse(zip_buffer, media_type="application/zip", headers={"Content-Disposition": "attachment; filename=BonnaCloud_Export.zip"})

if os.path.isdir("dist"): app.mount("/assets", StaticFiles(directory="dist/assets"), name="assets")
@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    if full_path.startswith("api/") or full_path == "files" or full_path.startswith("download/"): return JSONResponse(status_code=404, content={"detail": "API endpoint not found"})
    index_path = os.path.join("dist", "index.html")
    if os.path.exists(index_path): return FileResponse(index_path)
    return JSONResponse(status_code=404, content={"detail": "SPA Distribution Layer Missing"})